#!/usr/bin/env python3
"""
Excel Helper Script - Comprehensive utilities for Excel file manipulation
Provides common operations for creating, reading, editing, and formatting Excel files.

Dependencies:
    pip install openpyxl pandas

Author: Claude Code
Version: 1.0.0
"""

from typing import List, Dict, Any, Optional, Union, Tuple
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.chart import LineChart, BarChart, PieChart, ScatterChart, AreaChart, Reference
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook as WorkbookType

import pandas as pd


# ============================================================================
# WORKBOOK CREATION AND LOADING
# ============================================================================

def create_workbook(
    title: str = "Sheet1",
    headers: Optional[List[str]] = None,
    data: Optional[List[List[Any]]] = None
) -> Tuple[WorkbookType, Worksheet]:
    """
    Create a new Excel workbook with optional headers and data.

    Args:
        title: The worksheet title
        headers: List of column headers
        data: List of data rows (list of lists)

    Returns:
        Tuple of (Workbook, Worksheet)

    Example:
        >>> wb, ws = create_workbook("Sales", ["Product", "Q1", "Q2"])
        >>> wb.save("sales.xlsx")
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    if headers:
        ws.append(headers)
        # Format headers
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    if data:
        for row in data:
            ws.append(row)

    return wb, ws


def load_excel(
    file_path: str,
    sheet_name: Optional[str] = None,
    read_only: bool = False,
    data_only: bool = False
) -> Tuple[WorkbookType, Worksheet]:
    """
    Load an existing Excel workbook.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to load (default: active sheet)
        read_only: Open in read-only mode for performance
        data_only: Read calculated formula values instead of formulas

    Returns:
        Tuple of (Workbook, Worksheet)

    Example:
        >>> wb, ws = load_excel("data.xlsx", sheet_name="Sales")
        >>> print(ws['A1'].value)
    """
    if not Path(file_path).exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, read_only=read_only, data_only=data_only)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    return wb, ws


# ============================================================================
# DATA READING AND WRITING
# ============================================================================

def read_excel_data(
    file_path: str,
    sheet_name: Optional[str] = None,
    min_row: int = 1,
    max_row: Optional[int] = None,
    min_col: int = 1,
    max_col: Optional[int] = None,
    values_only: bool = True
) -> List[Tuple]:
    """
    Read data from an Excel file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to read from
        min_row: Starting row (1-indexed)
        max_row: Ending row (inclusive)
        min_col: Starting column (1-indexed)
        max_col: Ending column (inclusive)
        values_only: Return only values (True) or cell objects (False)

    Returns:
        List of tuples containing row data

    Example:
        >>> data = read_excel_data("sales.xlsx", min_row=2)
        >>> for row in data:
        ...     print(row)
    """
    wb, ws = load_excel(file_path, sheet_name, read_only=True, data_only=True)

    data = list(ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=values_only
    ))

    wb.close()
    return data


def write_data_to_sheet(
    ws: Worksheet,
    data: List[List[Any]],
    start_row: int = 1,
    start_col: int = 1
) -> None:
    """
    Write data to a worksheet starting at specified position.

    Args:
        ws: Worksheet object
        data: List of data rows (list of lists)
        start_row: Starting row (1-indexed)
        start_col: Starting column (1-indexed)

    Example:
        >>> data = [["Product A", 100], ["Product B", 200]]
        >>> write_data_to_sheet(ws, data, start_row=2)
    """
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            ws.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


def read_excel_to_dataframe(
    file_path: str,
    sheet_name: Optional[str] = None,
    header: Union[int, None] = 0
) -> pd.DataFrame:
    """
    Read Excel file into a pandas DataFrame.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to read (default: first sheet)
        header: Row number to use as column names (0-indexed)

    Returns:
        pandas DataFrame

    Example:
        >>> df = read_excel_to_dataframe("sales.xlsx")
        >>> print(df.head())
    """
    return pd.read_excel(file_path, sheet_name=sheet_name, header=header)


def write_dataframe_to_excel(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str = "Sheet1",
    index: bool = False,
    header: bool = True,
    format_headers: bool = True
) -> None:
    """
    Write pandas DataFrame to Excel file with optional header formatting.

    Args:
        df: pandas DataFrame to write
        file_path: Output file path
        sheet_name: Name of worksheet
        index: Write DataFrame index
        header: Write column names
        format_headers: Apply formatting to header row

    Example:
        >>> df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
        >>> write_dataframe_to_excel(df, "output.xlsx")
    """
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index, header=header)

        if format_headers and header:
            wb = writer.book
            ws = wb[sheet_name]

            # Format header row
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")


# ============================================================================
# FORMULA OPERATIONS
# ============================================================================

def add_formula(
    ws: Worksheet,
    cell: str,
    formula: str,
    copy_down: Optional[int] = None
) -> None:
    """
    Add a formula to a cell and optionally copy it down.

    Args:
        ws: Worksheet object
        cell: Cell reference (e.g., "A1")
        formula: Excel formula (with = sign)
        copy_down: Number of rows to copy formula down

    Example:
        >>> add_formula(ws, "D2", "=SUM(A2:C2)", copy_down=10)
    """
    if not formula.startswith('='):
        formula = '=' + formula

    ws[cell] = formula

    if copy_down:
        # Extract column and row
        col_letter = ''.join(filter(str.isalpha, cell))
        start_row = int(''.join(filter(str.isdigit, cell)))

        # Copy formula down
        for row_offset in range(1, copy_down):
            target_cell = f"{col_letter}{start_row + row_offset}"
            # Formula will automatically adjust relative references
            ws[target_cell] = formula.replace(
                str(start_row),
                str(start_row + row_offset)
            )


def add_sum_formula(
    ws: Worksheet,
    target_cell: str,
    range_start: str,
    range_end: str
) -> None:
    """
    Add a SUM formula to a cell.

    Args:
        ws: Worksheet object
        target_cell: Cell where formula will be placed
        range_start: Start of range to sum
        range_end: End of range to sum

    Example:
        >>> add_sum_formula(ws, "E10", "E2", "E9")
        >>> # Creates: =SUM(E2:E9)
    """
    formula = f"=SUM({range_start}:{range_end})"
    ws[target_cell] = formula


def add_average_formula(
    ws: Worksheet,
    target_cell: str,
    range_start: str,
    range_end: str
) -> None:
    """
    Add an AVERAGE formula to a cell.

    Args:
        ws: Worksheet object
        target_cell: Cell where formula will be placed
        range_start: Start of range to average
        range_end: End of range to average

    Example:
        >>> add_average_formula(ws, "F10", "F2", "F9")
    """
    formula = f"=AVERAGE({range_start}:{range_end})"
    ws[target_cell] = formula


# ============================================================================
# FORMATTING
# ============================================================================

def apply_formatting(
    ws: Worksheet,
    cell_range: str,
    bold: bool = False,
    italic: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border: bool = False,
    align: Optional[str] = None,
    number_format: Optional[str] = None
) -> None:
    """
    Apply formatting to a range of cells.

    Args:
        ws: Worksheet object
        cell_range: Range to format (e.g., "A1:C10")
        bold: Apply bold font
        italic: Apply italic font
        font_size: Font size in points
        font_color: Hex color code for font (e.g., "FF0000")
        bg_color: Hex color code for background (e.g., "FFFF00")
        border: Apply thin borders
        align: Alignment ("left", "center", "right")
        number_format: Number format string

    Example:
        >>> apply_formatting(ws, "A1:D1", bold=True, bg_color="4472C4")
    """
    # Create style objects
    font = Font(
        bold=bold,
        italic=italic,
        size=font_size,
        color=font_color
    )

    fill = None
    if bg_color:
        fill = PatternFill(
            start_color=bg_color,
            end_color=bg_color,
            fill_type="solid"
        )

    border_style = None
    if border:
        side = Side(style='thin')
        border_style = Border(left=side, right=side, top=side, bottom=side)

    alignment = None
    if align:
        alignment = Alignment(horizontal=align)

    # Apply to range
    for row in ws[cell_range]:
        for cell in row:
            if bold or italic or font_size or font_color:
                cell.font = font
            if fill:
                cell.fill = fill
            if border_style:
                cell.border = border_style
            if alignment:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format


def format_as_currency(
    ws: Worksheet,
    cell_range: str,
    currency: str = "USD"
) -> None:
    """
    Format cells as currency.

    Args:
        ws: Worksheet object
        cell_range: Range to format (e.g., "B2:B10")
        currency: Currency code ("USD", "EUR", "GBP", etc.)

    Example:
        >>> format_as_currency(ws, "C2:C10")
    """
    format_dict = {
        "USD": numbers.FORMAT_CURRENCY_USD_SIMPLE,
        "EUR": "[$€-407]#,##0.00",
        "GBP": "[$£-809]#,##0.00"
    }

    number_format = format_dict.get(currency, numbers.FORMAT_CURRENCY_USD_SIMPLE)

    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = number_format


def format_as_percentage(
    ws: Worksheet,
    cell_range: str,
    decimals: int = 2
) -> None:
    """
    Format cells as percentage.

    Args:
        ws: Worksheet object
        cell_range: Range to format
        decimals: Number of decimal places

    Example:
        >>> format_as_percentage(ws, "D2:D10", decimals=1)
    """
    if decimals == 0:
        format_str = "0%"
    else:
        format_str = f"0.{'0' * decimals}%"

    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = format_str


def format_as_date(
    ws: Worksheet,
    cell_range: str,
    date_format: str = "mm/dd/yyyy"
) -> None:
    """
    Format cells as dates.

    Args:
        ws: Worksheet object
        cell_range: Range to format
        date_format: Date format string

    Example:
        >>> format_as_date(ws, "E2:E10", "dd/mm/yyyy")
    """
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = date_format


def auto_fit_columns(
    ws: Worksheet,
    min_width: int = 8,
    max_width: int = 50
) -> None:
    """
    Auto-fit column widths based on content.

    Args:
        ws: Worksheet object
        min_width: Minimum column width
        max_width: Maximum column width

    Example:
        >>> auto_fit_columns(ws)
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def merge_cells_with_value(
    ws: Worksheet,
    cell_range: str,
    value: Any,
    bold: bool = True,
    align: str = "center"
) -> None:
    """
    Merge cells and set value with formatting.

    Args:
        ws: Worksheet object
        cell_range: Range to merge (e.g., "A1:D1")
        value: Value to set in merged cell
        bold: Apply bold font
        align: Text alignment

    Example:
        >>> merge_cells_with_value(ws, "A1:E1", "Sales Report 2024")
    """
    ws.merge_cells(cell_range)

    # Get top-left cell
    start_cell = cell_range.split(':')[0]
    cell = ws[start_cell]

    cell.value = value
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")


# ============================================================================
# CHART CREATION
# ============================================================================

def add_chart(
    ws: Worksheet,
    chart_type: str,
    data_range: str,
    categories_range: Optional[str] = None,
    title: Optional[str] = None,
    x_axis_title: Optional[str] = None,
    y_axis_title: Optional[str] = None,
    position: str = "F2",
    width: int = 15,
    height: int = 10
) -> None:
    """
    Add a chart to the worksheet.

    Args:
        ws: Worksheet object
        chart_type: Type of chart ("line", "bar", "pie", "scatter", "area")
        data_range: Range containing data (e.g., "B2:D10")
        categories_range: Range for x-axis labels (e.g., "A2:A10")
        title: Chart title
        x_axis_title: X-axis label
        y_axis_title: Y-axis label
        position: Cell where chart will be placed
        width: Chart width in Excel units
        height: Chart height in Excel units

    Example:
        >>> add_chart(ws, "line", "B2:D10", "A2:A10", "Sales Trend")
    """
    chart_classes = {
        "line": LineChart,
        "bar": BarChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart
    }

    if chart_type not in chart_classes:
        raise ValueError(f"Unsupported chart type: {chart_type}")

    chart = chart_classes[chart_type]()

    if title:
        chart.title = title

    if x_axis_title and hasattr(chart, 'x_axis'):
        chart.x_axis.title = x_axis_title

    if y_axis_title and hasattr(chart, 'y_axis'):
        chart.y_axis.title = y_axis_title

    chart.width = width
    chart.height = height

    # Parse data range - handle both with and without sheet name
    if '!' in data_range:
        data_ref = Reference(ws, range_string=data_range)
    else:
        # Parse range like "B2:D10"
        parts = data_range.split(':')
        min_cell = parts[0]
        max_cell = parts[1] if len(parts) > 1 else parts[0]

        # Extract row and column
        min_col = column_index_from_string(''.join(filter(str.isalpha, min_cell)))
        min_row = int(''.join(filter(str.isdigit, min_cell)))
        max_col = column_index_from_string(''.join(filter(str.isalpha, max_cell)))
        max_row = int(''.join(filter(str.isdigit, max_cell)))

        data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

    if chart_type == "pie":
        chart.add_data(data_ref)
        if categories_range:
            if '!' in categories_range:
                cat_ref = Reference(ws, range_string=categories_range)
            else:
                parts = categories_range.split(':')
                min_cell = parts[0]
                max_cell = parts[1] if len(parts) > 1 else parts[0]
                min_col = column_index_from_string(''.join(filter(str.isalpha, min_cell)))
                min_row = int(''.join(filter(str.isdigit, min_cell)))
                max_col = column_index_from_string(''.join(filter(str.isalpha, max_cell)))
                max_row = int(''.join(filter(str.isdigit, max_cell)))
                cat_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            chart.set_categories(cat_ref)
    else:
        chart.add_data(data_ref, titles_from_data=True)
        if categories_range:
            if '!' in categories_range:
                cat_ref = Reference(ws, range_string=categories_range)
            else:
                parts = categories_range.split(':')
                min_cell = parts[0]
                max_cell = parts[1] if len(parts) > 1 else parts[0]
                min_col = column_index_from_string(''.join(filter(str.isalpha, min_cell)))
                min_row = int(''.join(filter(str.isdigit, min_cell)))
                max_col = column_index_from_string(''.join(filter(str.isalpha, max_cell)))
                max_row = int(''.join(filter(str.isdigit, max_cell)))
                cat_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            chart.set_categories(cat_ref)

    ws.add_chart(chart, position)


def create_line_chart(
    ws: Worksheet,
    data_cols: List[int],
    category_col: int,
    start_row: int,
    end_row: int,
    title: str = "Line Chart",
    position: str = "F2"
) -> None:
    """
    Create a line chart with multiple series.

    Args:
        ws: Worksheet object
        data_cols: List of column numbers containing data series
        category_col: Column number for x-axis categories
        start_row: Starting row (inclusive)
        end_row: Ending row (inclusive)
        title: Chart title
        position: Cell position for chart

    Example:
        >>> create_line_chart(ws, [2, 3, 4], 1, 2, 10, "Monthly Sales")
    """
    chart = LineChart()
    chart.title = title
    chart.style = 10

    data = Reference(
        ws,
        min_col=min(data_cols),
        max_col=max(data_cols),
        min_row=start_row,
        max_row=end_row
    )

    categories = Reference(
        ws,
        min_col=category_col,
        min_row=start_row + 1,
        max_row=end_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, position)


# ============================================================================
# CONDITIONAL FORMATTING
# ============================================================================

def add_color_scale(
    ws: Worksheet,
    cell_range: str,
    start_color: str = "F8696B",
    mid_color: str = "FFEB84",
    end_color: str = "63BE7B"
) -> None:
    """
    Add color scale conditional formatting.

    Args:
        ws: Worksheet object
        cell_range: Range to format (e.g., "B2:B10")
        start_color: Hex color for minimum values
        mid_color: Hex color for middle values
        end_color: Hex color for maximum values

    Example:
        >>> add_color_scale(ws, "C2:C10")
    """
    from openpyxl.formatting.rule import ColorScaleRule

    rule = ColorScaleRule(
        start_type='min',
        start_color=start_color,
        mid_type='percentile',
        mid_value=50,
        mid_color=mid_color,
        end_type='max',
        end_color=end_color
    )

    ws.conditional_formatting.add(cell_range, rule)


def add_data_bars(
    ws: Worksheet,
    cell_range: str,
    color: str = "638EC6"
) -> None:
    """
    Add data bar conditional formatting.

    Args:
        ws: Worksheet object
        cell_range: Range to format
        color: Hex color for data bars

    Example:
        >>> add_data_bars(ws, "D2:D10", "FF0000")
    """
    from openpyxl.formatting.rule import DataBarRule

    rule = DataBarRule(
        start_type='min',
        end_type='max',
        color=color
    )

    ws.conditional_formatting.add(cell_range, rule)


def highlight_cells_above(
    ws: Worksheet,
    cell_range: str,
    threshold: float,
    color: str = "C6EFCE"
) -> None:
    """
    Highlight cells with values above a threshold.

    Args:
        ws: Worksheet object
        cell_range: Range to format
        threshold: Value threshold
        color: Hex color for highlight

    Example:
        >>> highlight_cells_above(ws, "E2:E10", 1000, "90EE90")
    """
    from openpyxl.formatting.rule import CellIsRule

    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    rule = CellIsRule(
        operator='greaterThan',
        formula=[threshold],
        fill=fill
    )

    ws.conditional_formatting.add(cell_range, rule)


def highlight_cells_below(
    ws: Worksheet,
    cell_range: str,
    threshold: float,
    color: str = "FFC7CE"
) -> None:
    """
    Highlight cells with values below a threshold.

    Args:
        ws: Worksheet object
        cell_range: Range to format
        threshold: Value threshold
        color: Hex color for highlight

    Example:
        >>> highlight_cells_below(ws, "F2:F10", 500, "FFB6C1")
    """
    from openpyxl.formatting.rule import CellIsRule

    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    rule = CellIsRule(
        operator='lessThan',
        formula=[threshold],
        fill=fill
    )

    ws.conditional_formatting.add(cell_range, rule)


# ============================================================================
# DATA VALIDATION
# ============================================================================

def add_dropdown_validation(
    ws: Worksheet,
    cell_range: str,
    options: List[str],
    error_message: str = "Please select from the list"
) -> None:
    """
    Add dropdown list data validation.

    Args:
        ws: Worksheet object
        cell_range: Range where validation applies
        options: List of dropdown options
        error_message: Error message for invalid input

    Example:
        >>> add_dropdown_validation(ws, "C2:C100", ["Active", "Inactive"])
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    # Join options with comma
    formula = f'"{",".join(options)}"'

    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False
    )

    dv.error = error_message
    dv.errorTitle = 'Invalid Entry'

    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_numeric_validation(
    ws: Worksheet,
    cell_range: str,
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
    error_message: Optional[str] = None
) -> None:
    """
    Add numeric range validation.

    Args:
        ws: Worksheet object
        cell_range: Range where validation applies
        min_value: Minimum allowed value
        max_value: Maximum allowed value
        error_message: Error message for invalid input

    Example:
        >>> add_numeric_validation(ws, "D2:D100", min_value=0, max_value=100)
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=min_value,
        formula2=max_value,
        allow_blank=True
    )

    if error_message:
        dv.error = error_message
    else:
        dv.error = f'Value must be between {min_value} and {max_value}'

    dv.errorTitle = 'Invalid Value'

    ws.add_data_validation(dv)
    dv.add(cell_range)


# ============================================================================
# WORKSHEET MANAGEMENT
# ============================================================================

def create_sheet(
    wb: WorkbookType,
    sheet_name: str,
    position: Optional[int] = None
) -> Worksheet:
    """
    Create a new worksheet in the workbook.

    Args:
        wb: Workbook object
        sheet_name: Name for new sheet
        position: Position index (0-based), None for end

    Returns:
        New Worksheet object

    Example:
        >>> ws = create_sheet(wb, "Q1 Sales", position=0)
    """
    if position is not None:
        ws = wb.create_sheet(sheet_name, position)
    else:
        ws = wb.create_sheet(sheet_name)

    return ws


def copy_sheet(
    wb: WorkbookType,
    source_sheet: str,
    new_sheet_name: str
) -> Worksheet:
    """
    Copy an existing worksheet.

    Args:
        wb: Workbook object
        source_sheet: Name of sheet to copy
        new_sheet_name: Name for the copy

    Returns:
        New Worksheet object

    Example:
        >>> ws = copy_sheet(wb, "Template", "Q1 Report")
    """
    source = wb[source_sheet]
    target = wb.copy_worksheet(source)
    target.title = new_sheet_name

    return target


def delete_sheet(wb: WorkbookType, sheet_name: str) -> None:
    """
    Delete a worksheet from the workbook.

    Args:
        wb: Workbook object
        sheet_name: Name of sheet to delete

    Example:
        >>> delete_sheet(wb, "Old Data")
    """
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def get_cell_range(
    start_cell: str,
    end_cell: str
) -> str:
    """
    Create a cell range string.

    Args:
        start_cell: Starting cell (e.g., "A1")
        end_cell: Ending cell (e.g., "D10")

    Returns:
        Range string (e.g., "A1:D10")

    Example:
        >>> range_str = get_cell_range("A1", "D10")
        >>> print(range_str)  # "A1:D10"
    """
    return f"{start_cell}:{end_cell}"


def column_letter_to_number(letter: str) -> int:
    """
    Convert column letter to number.

    Args:
        letter: Column letter (e.g., "A", "AA")

    Returns:
        Column number (1-indexed)

    Example:
        >>> col_num = column_letter_to_number("AA")
        >>> print(col_num)  # 27
    """
    return column_index_from_string(letter)


def column_number_to_letter(number: int) -> str:
    """
    Convert column number to letter.

    Args:
        number: Column number (1-indexed)

    Returns:
        Column letter

    Example:
        >>> letter = column_number_to_letter(27)
        >>> print(letter)  # "AA"
    """
    return get_column_letter(number)


def get_last_row(ws: Worksheet, column: Union[int, str] = 1) -> int:
    """
    Get the last row with data in a column.

    Args:
        ws: Worksheet object
        column: Column number (int) or letter (str)

    Returns:
        Last row number with data

    Example:
        >>> last_row = get_last_row(ws, "A")
        >>> print(f"Data up to row {last_row}")
    """
    if isinstance(column, str):
        column = column_letter_to_number(column)

    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=column).value is not None:
            return row

    return 0


def clear_range(ws: Worksheet, cell_range: str) -> None:
    """
    Clear values from a cell range.

    Args:
        ws: Worksheet object
        cell_range: Range to clear (e.g., "A1:D10")

    Example:
        >>> clear_range(ws, "B2:D10")
    """
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


# ============================================================================
# COMPLETE WORKFLOW EXAMPLES
# ============================================================================

def create_sales_report_template() -> None:
    """
    Create a complete sales report template with formatting and formulas.

    Example:
        >>> create_sales_report_template()
        >>> # Creates "sales_report_template.xlsx"
    """
    wb, ws = create_workbook("Sales Report")

    # Title
    merge_cells_with_value(ws, "A1:F1", "Monthly Sales Report", bold=True)
    ws.row_dimensions[1].height = 30
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Headers
    headers = ["Product", "Q1", "Q2", "Q3", "Q4", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    products = ["Product A", "Product B", "Product C", "Product D"]
    for row, product in enumerate(products, start=4):
        ws[f'A{row}'] = product
        ws[f'F{row}'] = f"=SUM(B{row}:E{row})"

    # Totals row
    total_row = len(products) + 4
    ws[f'A{total_row}'] = "Total"
    ws[f'A{total_row}'].font = Font(bold=True)

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{total_row}'] = f"=SUM({col}4:{col}{total_row-1})"
        ws[f'{col}{total_row}'].font = Font(bold=True)

    # Format as currency
    format_as_currency(ws, f"B4:F{total_row}")

    # Add data validation for numbers
    add_numeric_validation(ws, f"B4:E{total_row-1}", min_value=0)

    # Auto-fit columns
    auto_fit_columns(ws)

    wb.save("sales_report_template.xlsx")
    print("Sales report template created!")


def analyze_csv_and_create_report(csv_file: str, output_file: str) -> None:
    """
    Read CSV, analyze data, and create formatted Excel report with charts.

    Args:
        csv_file: Input CSV file path
        output_file: Output Excel file path

    Example:
        >>> analyze_csv_and_create_report("data.csv", "report.xlsx")
    """
    # Read CSV
    df = pd.read_csv(csv_file)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Raw Data sheet
    ws_data = wb.create_sheet("Raw Data")
    write_dataframe_to_excel(df, output_file, "Raw Data", format_headers=True)

    # Load back to add analysis
    wb = load_workbook(output_file)

    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)

    # Add summary statistics
    ws_summary['A1'] = "Data Summary"
    merge_cells_with_value(ws_summary, "A1:B1", "Data Summary", bold=True)

    stats = [
        ["Total Records", len(df)],
        ["Columns", len(df.columns)],
        ["Date Range", f"{df.iloc[0, 0]} to {df.iloc[-1, 0]}"]
    ]

    for row, (label, value) in enumerate(stats, start=2):
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)

    auto_fit_columns(ws_summary)

    wb.save(output_file)
    print(f"Analysis complete! Report saved to {output_file}")


# ============================================================================
# MAIN - DEMONSTRATION
# ============================================================================

def main():
    """
    Demonstration of helper functions.
    """
    print("Excel Helper Script - Comprehensive Utilities")
    print("=" * 60)

    # Example 1: Create simple workbook
    print("\n1. Creating simple workbook...")
    wb, ws = create_workbook(
        title="Demo",
        headers=["Name", "Score", "Grade"],
        data=[
            ["Alice", 95, "A"],
            ["Bob", 87, "B"],
            ["Charlie", 92, "A"]
        ]
    )

    # Add formulas
    add_average_formula(ws, "B5", "B2", "B4")
    ws['A5'] = "Average"
    ws['A5'].font = Font(bold=True)

    # Format
    apply_formatting(ws, "A1:C1", bold=True, bg_color="4472C4", font_color="FFFFFF")
    format_as_percentage(ws, "B2:B4", decimals=0)

    auto_fit_columns(ws)

    wb.save("demo_simple.xlsx")
    print("   ✓ Created demo_simple.xlsx")

    # Example 2: Create report with chart
    print("\n2. Creating report with chart...")
    create_sales_report_template()
    print("   ✓ Created sales_report_template.xlsx")

    print("\n" + "=" * 60)
    print("Demo complete! Check the generated files.")


if __name__ == "__main__":
    main()
